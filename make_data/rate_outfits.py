import random
import time
import tkinter as tk
from tkinter import messagebox
from pathlib import Path

from PIL import Image, ImageTk
from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_DIR / "data.xlsx"
ITEM_PICS_DIR = PROJECT_DIR / "item_pics"


def load_item_ids():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = wb["metadata"]
    tops, bottoms = [], []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_id, item_type = row[0], row[4]
        if item_id is None:
            continue
        if str(item_type).strip().lower() == "top":
            tops.append(int(item_id))
        else:
            bottoms.append(int(item_id))
    wb.close()
    return tops, bottoms


def append_score(top_id, bottom_id, temperature, score):
    wb = load_workbook(WORKBOOK_PATH)
    wb["data"].append([top_id, bottom_id, temperature, score])
    wb.save(WORKBOOK_PATH)
    wb.close()


def load_photo(item_id):
    img = Image.open(ITEM_PICS_DIR / f"{item_id}.png")
    w, h = img.size
    new_w = 300
    img = img.resize((new_w, int(h * new_w / w)), Image.LANCZOS)
    return ImageTk.PhotoImage(img)


class OutfitRaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Outfit Rater")

        self.tops, self.bottoms = load_item_ids()
        self.top_id = self.bottom_id = self.temperature = None
        self.top_photo = self.bottom_photo = None

        self.info_var = tk.StringVar()
        self.status_var = tk.StringVar()

        tk.Label(root, text="Rate Random Outfit Combinations", font=("Helvetica", 18, "bold")).pack(pady=8)
        tk.Label(root, textvariable=self.info_var, font=("Helvetica", 12)).pack()

        img_frame = tk.Frame(root)
        img_frame.pack(pady=8)
        tk.Label(img_frame, text="Top", font=("Helvetica", 14, "bold")).grid(row=0, column=0, padx=20)
        tk.Label(img_frame, text="Bottom", font=("Helvetica", 14, "bold")).grid(row=0, column=1, padx=20)
        self.top_label = tk.Label(img_frame)
        self.top_label.grid(row=1, column=0, padx=20)
        self.bottom_label = tk.Label(img_frame)
        self.bottom_label.grid(row=1, column=1, padx=20)

        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=8)
        for i in range(1, 11):
            text = str(i) if i < 10 else "10 (0)"
            tk.Button(btn_frame, text=text, width=6,
                      command=lambda s=i: self.save_score(s)).grid(row=0, column=i - 1, padx=3)

        footer = tk.Frame(root)
        footer.pack(fill="x", padx=16, pady=8)
        tk.Label(footer, textvariable=self.status_var, fg="#1f3b4d").pack(side="left")
        tk.Button(footer, text="Skip / New Random Outfit", command=self.next_outfit).pack(side="right")

        for i in range(1, 10):
            root.bind(f"<KeyPress-{i}>", lambda e, s=i: self.save_score(s))
        root.bind("<KeyPress-0>", lambda e: self.save_score(10))
        root.bind("<space>", lambda e: self.next_outfit())

        self.next_outfit()

    def next_outfit(self, status=None):
        self.top_id = random.choice(self.tops)
        self.bottom_id = random.choice(self.bottoms)
        self.temperature = random.randint(20, 90)

        try:
            self.top_photo = load_photo(self.top_id)
            self.bottom_photo = load_photo(self.bottom_id)
        except Exception as e:
            messagebox.showerror("Image Error", str(e))
            return

        self.top_label.configure(image=self.top_photo)
        self.bottom_label.configure(image=self.bottom_photo)
        self.info_var.set(f"Top #{self.top_id}    Bottom #{self.bottom_id}    Temp: {self.temperature}°F")
        self.status_var.set(status or "Press 1–9 or 0 (=10) to score.")

    def save_score(self, score):
        if self.top_id is None:
            return
        try:
            append_score(self.top_id, self.bottom_id, self.temperature, score)
        except Exception as e:
            messagebox.showerror("Save Error", str(e))
            return
        self.next_outfit(f"Saved: top #{self.top_id}, bottom #{self.bottom_id}, {self.temperature}°F, score {score}.")


if __name__ == "__main__":
    random.seed(time.time_ns())
    root = tk.Tk()
    OutfitRaterApp(root)
    root.mainloop()
